import csv
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook


WORKFLOW_CREATE_STUDENTS = "create_students"
WORKFLOW_CREATE_AND_MAP = "create_and_map"
WORKFLOW_MAP_EXISTING = "map_existing"
SUPPORTED_BULK_UPLOAD_EXTENSIONS = {".csv", ".xlsx"}


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_email(value: Any) -> str:
    return normalize_text(value).lower()


def status_rank(status_value: str) -> int:
    return {
        "invalid": 0,
        "blocked": 1,
        "valid": 2,
        "noop": 3,
        "failed": 4,
        "committed": 5,
    }.get(status_value, 99)


def row_identifier(row: dict[str, Any]) -> str:
    for key in ("email", "enrollment_number", "student_id", "roll_number", "full_name"):
        value = normalize_text(row.get(key))
        if value:
            return value
    return f"row-{row.get('row_number', '?')}"


def student_identifier_candidates(student: dict[str, Any]) -> list[str]:
    candidates = [str(student.get("_id"))]
    roll_number = normalize_text(student.get("roll_number"))
    enrollment_number = normalize_text(student.get("enrollment_number"))
    email = normalize_email(student.get("email"))
    if roll_number:
        candidates.append(roll_number)
    if enrollment_number:
        candidates.append(enrollment_number)
    if email:
        candidates.append(email)
    return list(dict.fromkeys(candidates))


def compact_row_audit(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "row_number": row.get("row_number"),
        "status": row.get("status"),
        "action": row.get("action"),
        "identifier": row.get("identifier"),
        "student_id": row.get("student_id"),
        "messages": row.get("messages"),
    }


def canonicalize_workflow(workflow: str) -> str:
    normalized = normalize_text(workflow)
    if normalized == WORKFLOW_CREATE_AND_MAP:
        return WORKFLOW_CREATE_STUDENTS
    return normalized


def is_create_students_workflow(workflow: str) -> bool:
    return canonicalize_workflow(workflow) == WORKFLOW_CREATE_STUDENTS


def _headers_for_workflow(workflow: str) -> set[str]:
    if is_create_students_workflow(workflow):
        return {"full_name", "email", "roll_number", "enrollment_number", "phone"}
    return {"student_id", "enrollment_number", "email", "group"}


async def parse_bulk_upload_rows(file: UploadFile, workflow: str) -> list[dict[str, str]]:
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in SUPPORTED_BULK_UPLOAD_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Upload a CSV or XLSX file.",
        )

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded file is empty.")

    rows: list[dict[str, str]] = []
    if suffix == ".csv":
        decoded = raw.decode("utf-8-sig")
        reader = csv.DictReader(StringIO(decoded))
        headers = {normalize_text(header).lower() for header in (reader.fieldnames or []) if header}
        if not headers.intersection(_headers_for_workflow(workflow)):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Upload template columns are missing.")
        for index, record in enumerate(reader, start=2):
            rows.append(
                {
                    "row_number": index,
                    **{normalize_text(key).lower(): normalize_text(value) for key, value in (record or {}).items() if key},
                }
            )
    else:
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
        worksheet = workbook.active
        raw_headers = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [normalize_text(value).lower() for value in (raw_headers or [])]
        if not set(headers).intersection(_headers_for_workflow(workflow)):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Upload template columns are missing.")
        for index, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            row: dict[str, str] = {"row_number": index}
            has_content = False
            for header, value in zip(headers, values or []):
                if not header:
                    continue
                normalized = normalize_text(value)
                row[header] = normalized
                if normalized:
                    has_content = True
            if has_content:
                rows.append(row)

    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Upload file contains no data rows.")
    return rows
