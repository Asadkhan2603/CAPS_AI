from __future__ import annotations

import argparse
import asyncio
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
import json
from pathlib import Path
import re
import sys
from typing import Any

from openpyxl import load_workbook

BACKEND_ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = BACKEND_ROOT.parent
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.core.database import db
from app.core.indexes import ensure_indexes
from app.core.schema_versions import (
    DEPARTMENT_SCHEMA_VERSION,
    FACULTY_SCHEMA_VERSION,
    PROGRAM_SCHEMA_VERSION,
    SPECIALIZATION_SCHEMA_VERSION,
    UNIVERSITY_SCHEMA_VERSION,
)
from app.services.academic_hierarchy import validate_duration_and_semesters
from app.services.master_hierarchy import (
    CANONICAL_MASTER_FIELD_CONTRACT,
    build_department_business_id,
    build_faculty_business_id,
    build_program_business_id,
    build_specialization_business_id,
    coalesce_code,
    normalize_code,
    normalize_text,
    validate_business_identifier,
)

WORKBOOK_PATH = REPO_ROOT / "exports" / "Master_copy.xlsx"
SUMMARY_PATH = REPO_ROOT / "docs" / "migrations" / "MASTER_HIERARCHY_IMPORT_SUMMARY.md"
EXPORTS_ROOT = REPO_ROOT / "exports"
MASTER_COLLECTIONS = ("specializations", "programs", "departments", "faculties", "universities")
MASTER_PAYLOAD_ATTRS = {
    "universities": "universities",
    "faculties": "faculties",
    "departments": "departments",
    "programs": "programs",
    "specializations": "specializations",
}
MASTER_ID_FIELDS = {
    "universities": "university_id",
    "faculties": "faculty_id",
    "departments": "department_id",
    "programs": "program_id",
    "specializations": "specialization_id",
}
DOWNSTREAM_MASTER_REFERENCE_FIELDS: dict[str, tuple[tuple[str, str], ...]] = {
    "faculties": (("users", "faculty_id"), ("classes", "faculty_id")),
    "departments": (("users", "department_id"), ("classes", "department_id")),
    "programs": (
        ("users", "program_id"),
        ("batches", "program_id"),
        ("semesters", "program_id"),
        ("classes", "program_id"),
    ),
    "specializations": (
        ("users", "specialization_id"),
        ("batches", "specialization_id"),
        ("semesters", "specialization_id"),
        ("classes", "specialization_id"),
    ),
}

NORMALIZED_HEADER_MAP = {
    "department_id": "department_id",
    "department_name": "department_name",
    "department_code": "department_code",
    "program_id": "program_id",
    "program_code": "program_code",
    "program_name": "program_name",
    "durationyears": "duration_years",
    "total_semesters": "total_semesters",
    "degree_type": "degree_type",
    "specialization_id": "specialization_id",
    "specialization_code": "specialization_code",
    "specialization": "specialization_name",
    "faculty_id": "faculty_id",
    "faculty_code": "faculty_code",
    "faculty_name": "faculty_name",
    "university_id": "university_id",
    "university_name": "university_name",
    "department": "department_name",
}


@dataclass
class WorkbookPayload:
    universities: list[dict[str, Any]]
    faculties: list[dict[str, Any]]
    departments: list[dict[str, Any]]
    programs: list[dict[str, Any]]
    specializations: list[dict[str, Any]]
    rough_cross_check: list[dict[str, Any]]
    generic_patterns: set[str]


def _json_safe(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value) if hasattr(value, "binary") else value


def normalize_header(value: Any) -> str | None:
    text = normalize_text(value)
    if not text:
        return None
    text = text.replace("(", "_").replace(")", "")
    text = re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_").lower()
    return NORMALIZED_HEADER_MAP.get(text, text)


def normalize_cell(value: Any) -> Any:
    if isinstance(value, str):
        text = value.strip()
        return text or None
    return value


def load_sheet_rows(workbook, sheet_name: str) -> list[dict[str, Any]]:
    ws = workbook[sheet_name]
    headers = [normalize_header(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = {header: normalize_cell(value) for header, value in zip(headers, row) if header}
        if not any(value not in (None, "") for value in values.values()):
            continue
        rows.append(values)
    return rows


def load_generic_patterns(workbook) -> set[str]:
    ws = workbook["generic_format"]
    patterns: set[str] = set()
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            text = normalize_text(cell)
            if not text:
                continue
            if text.startswith(("FAC-", "DEP-", "PRG-", "SPC-")):
                patterns.add(text)
    return patterns


def load_workbook_payload(path: Path) -> WorkbookPayload:
    workbook = load_workbook(path, data_only=True)
    required_sheets = {
        "University_details",
        "FACULTY_DATA",
        "DEPARTMENT_DATA",
        "PROGRAM_DATA",
        "SPECIALIZATION_DATA",
        "generic_format",
        "rough_sheet",
    }
    missing = sorted(required_sheets.difference(workbook.sheetnames))
    if missing:
        raise ValueError(f"Workbook is missing required sheets: {', '.join(missing)}")
    return WorkbookPayload(
        universities=load_sheet_rows(workbook, "University_details"),
        faculties=load_sheet_rows(workbook, "FACULTY_DATA"),
        departments=load_sheet_rows(workbook, "DEPARTMENT_DATA"),
        programs=load_sheet_rows(workbook, "PROGRAM_DATA"),
        specializations=load_sheet_rows(workbook, "SPECIALIZATION_DATA"),
        rough_cross_check=load_sheet_rows(workbook, "rough_sheet"),
        generic_patterns=load_generic_patterns(workbook),
    )


def _ensure_unique(rows: list[dict[str, Any]], field: str, label: str) -> list[str]:
    counter = Counter(str(row.get(field) or "") for row in rows if row.get(field))
    return [f"Duplicate {label} '{value}' appears {count} times." for value, count in counter.items() if count > 1]


def _ensure_scoped_unique(rows: list[dict[str, Any]], scope_fields: tuple[str, ...], field: str, label: str) -> list[str]:
    counter = Counter(
        tuple(str(row.get(item) or "") for item in (*scope_fields, field))
        for row in rows
        if row.get(field)
    )
    messages = []
    for values, count in counter.items():
        if count <= 1:
            continue
        scope_values = ", ".join(f"{name}={value}" for name, value in zip((*scope_fields, field), values))
        messages.append(f"Duplicate {label} combination '{scope_values}' appears {count} times.")
    return messages


def validate_workbook_payload(payload: WorkbookPayload) -> dict[str, Any]:
    errors: list[str] = []
    reconciliations: list[dict[str, str]] = []
    expected_patterns = {
        "FAC-[FAC_CODE]",
        "DEP-[FAC_CODE]-[DEPT_CODE]",
        "PRG-[FAC_CODE]-[DEPT_CODE]-[PROGRAM_CODE]",
        "SPC-[FAC_CODE]-[DEPT_CODE]-[PROGRAM_CODE]-[SPEC_CODE]",
    }
    missing_patterns = sorted(expected_patterns.difference(payload.generic_patterns))
    if missing_patterns:
        errors.append("generic_format sheet is missing expected patterns: " + ", ".join(missing_patterns))

    university_by_business_id = {
        normalize_code(row["university_id"]): row
        for row in payload.universities
        if row.get("university_id")
    }
    faculty_by_business_id = {
        normalize_code(row["faculty_id"]): row
        for row in payload.faculties
        if row.get("faculty_id")
    }
    department_by_business_id = {
        normalize_code(row["department_id"]): row
        for row in payload.departments
        if row.get("department_id")
    }
    program_by_department_and_code = {
        (normalize_code(row.get("department_id")), normalize_code(row.get("program_code"))): row
        for row in payload.programs
        if row.get("department_id") and row.get("program_code")
    }

    for university in payload.universities:
        if not normalize_code(university.get("university_id")) or not normalize_text(university.get("university_name")):
            errors.append(f"University row is missing required values: {university}")

    for faculty in payload.faculties:
        university_id = normalize_code(faculty.get("university_id"))
        faculty_id = normalize_code(faculty.get("faculty_id"))
        faculty_code = coalesce_code(faculty.get("faculty_code"))
        faculty_name = normalize_text(faculty.get("faculty_name"))
        if not university_id or university_id not in university_by_business_id:
            errors.append(f"Faculty '{faculty_id or faculty_name or faculty}' references missing university_id '{university_id}'.")
            continue
        if not faculty_id or not faculty_code or not faculty_name:
            errors.append(f"Faculty row is missing required values: {faculty}")
            continue
        expected_faculty_id = build_faculty_business_id(faculty_code)
        if faculty_id != expected_faculty_id:
            reconciliations.append({"entity": "faculty", "name": faculty_name or faculty_id or "", "from": faculty_id or "", "to": expected_faculty_id})
            faculty["faculty_id"] = expected_faculty_id
            faculty_id = expected_faculty_id
        try:
            validate_business_identifier("faculty", faculty_id)
        except ValueError as exc:
            errors.append(str(exc))

    faculty_by_business_id = {
        normalize_code(row["faculty_id"]): row
        for row in payload.faculties
        if row.get("faculty_id")
    }

    for department in payload.departments:
        faculty_id = normalize_code(department.get("faculty_id"))
        faculty = faculty_by_business_id.get(faculty_id)
        department_id = normalize_code(department.get("department_id"))
        department_code = coalesce_code(department.get("department_code"))
        department_name = normalize_text(department.get("department_name"))
        provided_faculty_code = coalesce_code(department.get("faculty_code"))
        if not faculty:
            errors.append(f"Department '{department_id or department_name or department}' references missing faculty_id '{faculty_id}'.")
            continue
        if provided_faculty_code and provided_faculty_code != coalesce_code(faculty.get("faculty_code")):
            errors.append(
                f"Department '{department_name or department_id}' has faculty_code '{provided_faculty_code}' but parent faculty uses '{faculty.get('faculty_code')}'."
            )
        if not department_id or not department_code or not department_name:
            errors.append(f"Department row is missing required values: {department}")
            continue
        expected_department_id = build_department_business_id(
            faculty_code=str(faculty.get("faculty_code") or ""),
            department_code=department_code,
        )
        if department_id != expected_department_id:
            reconciliations.append({"entity": "department", "name": department_name or department_id or "", "from": department_id or "", "to": expected_department_id})
            department["department_id"] = expected_department_id
            department_id = expected_department_id
        try:
            validate_business_identifier("department", department_id)
        except ValueError as exc:
            errors.append(str(exc))

    department_by_business_id = {
        normalize_code(row["department_id"]): row
        for row in payload.departments
        if row.get("department_id")
    }

    for program in payload.programs:
        department_id = normalize_code(program.get("department_id"))
        department = department_by_business_id.get(department_id)
        program_id = normalize_code(program.get("program_id"))
        program_code = coalesce_code(program.get("program_code"))
        program_name = normalize_text(program.get("program_name"))
        provided_department_name = normalize_text(program.get("department_name"))
        if not department:
            errors.append(f"Program '{program_id or program_name or program}' references missing department_id '{department_id}'.")
            continue
        if provided_department_name and provided_department_name != normalize_text(department.get("department_name")):
            errors.append(
                f"Program '{program_name or program_id}' has department_name '{provided_department_name}' but parent department uses '{department.get('department_name')}'."
            )
        if not program_id or not program_code or not program_name:
            errors.append(f"Program row is missing required values: {program}")
            continue
        faculty = faculty_by_business_id.get(normalize_code(department.get("faculty_id")))
        expected_program_id = build_program_business_id(
            faculty_code=str((faculty or {}).get("faculty_code") or ""),
            department_code=str(department.get("department_code") or ""),
            program_code=program_code,
        )
        if program_id != expected_program_id:
            reconciliations.append({"entity": "program", "name": program_name or program_id or "", "from": program_id or "", "to": expected_program_id})
            program["program_id"] = expected_program_id
            program_id = expected_program_id
        try:
            validate_business_identifier("program", program_id)
        except ValueError as exc:
            errors.append(str(exc))
        try:
            validate_duration_and_semesters(program.get("duration_years"), program.get("total_semesters"))
        except ValueError as exc:
            errors.append(f"Program '{program_name}' has invalid duration/semester values: {exc}")

    program_by_department_and_code = {
        (normalize_code(row.get("department_id")), normalize_code(row.get("program_code"))): row
        for row in payload.programs
        if row.get("department_id") and row.get("program_code")
    }

    for specialization in payload.specializations:
        faculty_id = normalize_code(specialization.get("faculty_id"))
        department_id = normalize_code(specialization.get("department_id"))
        program_code = coalesce_code(specialization.get("program_code"))
        specialization_id = normalize_code(specialization.get("specialization_id"))
        specialization_code = coalesce_code(specialization.get("specialization_code"))
        specialization_name = normalize_text(specialization.get("specialization_name"))
        faculty = faculty_by_business_id.get(faculty_id)
        department = department_by_business_id.get(department_id)
        program = program_by_department_and_code.get((department_id, program_code))
        if not faculty or not department or not program:
            errors.append(
                "Specialization row has missing ancestry: "
                f"faculty_id={faculty_id}, department_id={department_id}, program_code={program_code}"
            )
            continue
        if coalesce_code(specialization.get("faculty_code")) and coalesce_code(specialization.get("faculty_code")) != coalesce_code(faculty.get("faculty_code")):
            errors.append(
                f"Specialization '{specialization_name or specialization_id}' has faculty_code '{specialization.get('faculty_code')}' but parent faculty uses '{faculty.get('faculty_code')}'."
            )
        if coalesce_code(specialization.get("department_code")) and coalesce_code(specialization.get("department_code")) != coalesce_code(department.get("department_code")):
            errors.append(
                f"Specialization '{specialization_name or specialization_id}' has department_code '{specialization.get('department_code')}' but parent department uses '{department.get('department_code')}'."
            )
        if normalize_text(specialization.get("department_name")) and normalize_text(specialization.get("department_name")) != normalize_text(department.get("department_name")):
            errors.append(
                f"Specialization '{specialization_name or specialization_id}' has department name '{specialization.get('department_name')}' but parent department uses '{department.get('department_name')}'."
            )
        if normalize_text(specialization.get("program_name")) and normalize_text(specialization.get("program_name")) != normalize_text(program.get("program_name")):
            errors.append(
                f"Specialization '{specialization_name or specialization_id}' has program name '{specialization.get('program_name')}' but parent program uses '{program.get('program_name')}'."
            )
        if not specialization_id or not specialization_code or not specialization_name:
            errors.append(f"Specialization row is missing required values: {specialization}")
            continue
        expected_specialization_id = build_specialization_business_id(
            faculty_code=str(faculty.get("faculty_code") or ""),
            department_code=str(department.get("department_code") or ""),
            program_code=str(program.get("program_code") or ""),
            specialization_code=specialization_code,
        )
        if specialization_id != expected_specialization_id:
            reconciliations.append({"entity": "specialization", "name": specialization_name or specialization_id or "", "from": specialization_id or "", "to": expected_specialization_id})
            specialization["specialization_id"] = expected_specialization_id
            specialization_id = expected_specialization_id
        try:
            validate_business_identifier("specialization", specialization_id)
        except ValueError as exc:
            errors.append(str(exc))

    errors.extend(_ensure_unique(payload.universities, "university_id", "university_id"))
    errors.extend(_ensure_unique(payload.faculties, "faculty_id", "faculty_id"))
    errors.extend(_ensure_unique(payload.faculties, "faculty_code", "faculty_code"))
    errors.extend(_ensure_unique(payload.departments, "department_id", "department_id"))
    errors.extend(_ensure_scoped_unique(payload.departments, ("faculty_id",), "department_code", "department_code"))
    errors.extend(_ensure_unique(payload.programs, "program_id", "program_id"))
    errors.extend(_ensure_scoped_unique(payload.programs, ("department_id",), "program_code", "program_code"))
    errors.extend(_ensure_unique(payload.specializations, "specialization_id", "specialization_id"))
    errors.extend(
        _ensure_scoped_unique(
            payload.specializations,
            ("department_id", "program_code"),
            "specialization_code",
            "specialization_code",
        )
    )

    rough_keys = set()
    for row in payload.rough_cross_check:
        faculty_id = normalize_code(row.get("faculty_id"))
        department_id = normalize_code(row.get("department_id"))
        faculty = faculty_by_business_id.get(faculty_id)
        department = department_by_business_id.get(department_id)
        if not faculty_id or not department_id:
            continue
        expected_program_id = normalize_code(row.get("program_id"))
        if department and row.get("program_code"):
            faculty = faculty_by_business_id.get(normalize_code(department.get("faculty_id"))) or faculty
            expected_program_id = build_program_business_id(
                faculty_code=str((faculty or {}).get("faculty_code") or ""),
                department_code=str(department.get("department_code") or ""),
                program_code=str(row.get("program_code") or ""),
            )
        rough_keys.add((faculty_id, department_id, expected_program_id))
    normalized_keys = {
        (normalize_code(department.get("faculty_id")), normalize_code(program.get("department_id")), normalize_code(program.get("program_id")))
        for program in payload.programs
        for department in [department_by_business_id.get(normalize_code(program.get("department_id")))]
        if department
    }
    if rough_keys.difference(normalized_keys):
        errors.append(
            f"rough_sheet contains {len(rough_keys.difference(normalized_keys))} faculty/department/program combinations not present in normalized sheets."
        )

    if errors:
        raise ValueError("\n".join(errors))

    return {
        "sheet_counts": {
            "universities": len(payload.universities),
            "faculties": len(payload.faculties),
            "departments": len(payload.departments),
            "programs": len(payload.programs),
            "specializations": len(payload.specializations),
        },
        "reconciliations": reconciliations,
    }


def _payload_rows_for_collection(payload: WorkbookPayload, collection_name: str) -> list[dict[str, Any]]:
    return getattr(payload, MASTER_PAYLOAD_ATTRS[collection_name])


def _normalize_state_row(collection_name: str, row: dict[str, Any]) -> dict[str, Any]:
    if collection_name == "universities":
        return {
            "university_id": row.get("university_id"),
            "university_name": row.get("university_name"),
        }
    if collection_name == "faculties":
        return {
            "faculty_id": row.get("faculty_id"),
            "faculty_code": row.get("faculty_code"),
            "faculty_name": row.get("faculty_name"),
            "parent_business_id": row.get("university_master_id") or row.get("university_id"),
        }
    if collection_name == "departments":
        return {
            "department_id": row.get("department_id"),
            "department_code": row.get("department_code"),
            "department_name": row.get("department_name"),
            "parent_business_id": row.get("faculty_master_id") or row.get("faculty_id"),
        }
    if collection_name == "programs":
        return {
            "program_id": row.get("program_id"),
            "program_code": row.get("program_code"),
            "program_name": row.get("program_name"),
            "parent_business_id": row.get("department_master_id") or row.get("department_id"),
            "duration_years": row.get("duration_years"),
            "total_semesters": row.get("total_semesters"),
            "degree_type": row.get("degree_type"),
        }
    return {
        "specialization_id": row.get("specialization_id"),
        "specialization_code": row.get("specialization_code"),
        "specialization_name": row.get("specialization_name"),
        "parent_business_id": row.get("program_master_id") or row.get("program_id"),
    }


def _normalize_payload_row(
    collection_name: str,
    row: dict[str, Any],
    *,
    payload_program_lookup: dict[tuple[str | None, str | None], str],
) -> dict[str, Any]:
    if collection_name == "universities":
        return {
            "university_id": row.get("university_id"),
            "university_name": row.get("university_name"),
        }
    if collection_name == "faculties":
        return {
            "faculty_id": row.get("faculty_id"),
            "faculty_code": row.get("faculty_code"),
            "faculty_name": row.get("faculty_name"),
            "parent_business_id": row.get("university_id"),
        }
    if collection_name == "departments":
        return {
            "department_id": row.get("department_id"),
            "department_code": row.get("department_code"),
            "department_name": row.get("department_name"),
            "parent_business_id": row.get("faculty_id"),
        }
    if collection_name == "programs":
        duration_years = row.get("duration_years")
        total_semesters = row.get("total_semesters")
        return {
            "program_id": row.get("program_id"),
            "program_code": row.get("program_code"),
            "program_name": row.get("program_name"),
            "parent_business_id": row.get("department_id"),
            "duration_years": int(duration_years) if duration_years is not None else None,
            "total_semesters": int(total_semesters) if total_semesters is not None else None,
            "degree_type": row.get("degree_type"),
        }
    return {
        "specialization_id": row.get("specialization_id"),
        "specialization_code": row.get("specialization_code"),
        "specialization_name": row.get("specialization_name"),
        "parent_business_id": payload_program_lookup.get(
            (normalize_code(row.get("department_id")), normalize_code(row.get("program_code")))
        ),
    }


def build_master_change_plan(
    *,
    current_state: dict[str, list[dict[str, Any]]],
    payload: WorkbookPayload,
) -> dict[str, Any]:
    plan: dict[str, Any] = {}
    payload_program_lookup = {
        (normalize_code(row.get("department_id")), normalize_code(row.get("program_code"))): str(row.get("program_id") or "")
        for row in payload.programs
    }
    for collection_name in MASTER_COLLECTIONS[::-1]:
        identifier = MASTER_ID_FIELDS[collection_name]
        current_rows = {
            normalize_code(row.get(identifier)): _normalize_state_row(collection_name, row)
            for row in current_state.get(collection_name, [])
            if row.get(identifier)
        }
        incoming_rows = {
            normalize_code(row.get(identifier)): _normalize_payload_row(
                collection_name,
                row,
                payload_program_lookup=payload_program_lookup,
            )
            for row in _payload_rows_for_collection(payload, collection_name)
            if row.get(identifier)
        }

        added_ids = sorted(set(incoming_rows).difference(current_rows))
        removed_ids = sorted(set(current_rows).difference(incoming_rows))
        updated_ids = sorted(
            item_id
            for item_id in set(current_rows).intersection(incoming_rows)
            if current_rows[item_id] != incoming_rows[item_id]
        )
        unchanged_count = len(set(current_rows).intersection(incoming_rows)) - len(updated_ids)
        plan[collection_name] = {
            "current_count": len(current_rows),
            "incoming_count": len(incoming_rows),
            "added": {"count": len(added_ids), "sample_ids": added_ids[:10]},
            "updated": {"count": len(updated_ids), "sample_ids": updated_ids[:10]},
            "removed": {"count": len(removed_ids), "sample_ids": removed_ids[:10]},
            "unchanged_count": unchanged_count,
        }
    return plan


def change_plan_has_mutations(change_plan: dict[str, Any]) -> bool:
    return any(
        plan["added"]["count"] or plan["updated"]["count"] or plan["removed"]["count"]
        for plan in change_plan.values()
    )


def format_downstream_blockers(blockers: list[dict[str, Any]]) -> str:
    if not blockers:
        return "No downstream blockers detected."
    lines = ["Downstream blockers detected:"]
    for blocker in blockers:
        lines.append(
            "- "
            f"{blocker['master_collection']} is still referenced by {blocker['count']} records in "
            f"{blocker['referenced_by_collection']}.{blocker['field']}"
        )
        for sample in blocker.get("sample") or []:
            sample_bits = ", ".join(f"{key}={value}" for key, value in sample.items())
            lines.append(f"  sample: {sample_bits}")
    return "\n".join(lines)


async def _existing_master_docs(collection_name: str) -> list[dict[str, Any]]:
    collection = getattr(db, collection_name)
    return await collection.find({}, {"_id": 1}).to_list(length=10000)


async def fetch_current_master_state() -> dict[str, list[dict[str, Any]]]:
    projections = {
        "universities": {"university_id": 1, "university_name": 1, "is_active": 1},
        "faculties": {
            "faculty_id": 1,
            "faculty_code": 1,
            "faculty_name": 1,
            "university_id": 1,
            "university_master_id": 1,
            "is_active": 1,
        },
        "departments": {
            "department_id": 1,
            "department_code": 1,
            "department_name": 1,
            "faculty_id": 1,
            "faculty_master_id": 1,
            "university_master_id": 1,
            "is_active": 1,
        },
        "programs": {
            "program_id": 1,
            "program_code": 1,
            "program_name": 1,
            "department_id": 1,
            "department_master_id": 1,
            "faculty_master_id": 1,
            "duration_years": 1,
            "total_semesters": 1,
            "degree_type": 1,
            "is_active": 1,
        },
        "specializations": {
            "specialization_id": 1,
            "specialization_code": 1,
            "specialization_name": 1,
            "program_id": 1,
            "program_master_id": 1,
            "department_master_id": 1,
            "faculty_master_id": 1,
            "is_active": 1,
        },
    }
    state: dict[str, list[dict[str, Any]]] = {}
    for collection_name in MASTER_COLLECTIONS[::-1]:
        projection = {"_id": 1, **projections[collection_name]}
        state[collection_name] = await getattr(db, collection_name).find({}, projection).to_list(length=10000)
    return state


def write_master_backup(
    *,
    current_state: dict[str, list[dict[str, Any]]],
    backup_root: Path | None = None,
) -> Path:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    target_dir = (backup_root or EXPORTS_ROOT) / f"master_hierarchy_backup_{timestamp}"
    target_dir.mkdir(parents=True, exist_ok=True)
    for collection_name, rows in current_state.items():
        serialized = [{key: _json_safe(value) for key, value in row.items()} for row in rows]
        (target_dir / f"{collection_name}.json").write_text(json.dumps(serialized, indent=2), encoding="utf-8")
    return target_dir


async def detect_downstream_blockers() -> list[dict[str, Any]]:
    blockers: list[dict[str, Any]] = []
    existing_ids_by_collection: dict[str, list[str]] = {}
    for collection_name in MASTER_COLLECTIONS:
        rows = await _existing_master_docs(collection_name)
        existing_ids_by_collection[collection_name] = [str(row["_id"]) for row in rows if row.get("_id")]

    for master_collection, refs in DOWNSTREAM_MASTER_REFERENCE_FIELDS.items():
        master_ids = existing_ids_by_collection.get(master_collection) or []
        if not master_ids:
            continue
        for collection_name, field_name in refs:
            collection = getattr(db, collection_name, None)
            if collection is None:
                continue
            count = await collection.count_documents({field_name: {"$in": master_ids}})
            if count:
                sample = await collection.find({field_name: {"$in": master_ids}}, {"_id": 1, field_name: 1}).limit(5).to_list(length=5)
                blockers.append(
                    {
                        "master_collection": master_collection,
                        "referenced_by_collection": collection_name,
                        "field": field_name,
                        "count": count,
                        "sample": [{key: str(value) for key, value in row.items()} for row in sample],
                    }
                )
    return blockers


async def replace_master_data(payload: WorkbookPayload) -> dict[str, Any]:
    blockers = await detect_downstream_blockers()
    if blockers:
        raise RuntimeError(
            "Cannot replace master hierarchy while downstream records still reference existing master documents:\n"
            + json.dumps(blockers, indent=2)
        )

    for collection_name in MASTER_COLLECTIONS:
        await getattr(db, collection_name).delete_many({})

    now = datetime.now(timezone.utc)
    university_doc_by_business_id: dict[str, dict[str, Any]] = {}
    for row in payload.universities:
        document = {
            "university_id": normalize_code(row["university_id"]),
            "university_name": normalize_text(row["university_name"]),
            "is_active": True,
            "created_at": now,
            "schema_version": UNIVERSITY_SCHEMA_VERSION,
        }
        result = await db.universities.insert_one(document)
        university_doc_by_business_id[document["university_id"]] = {**document, "_id": result.inserted_id}

    faculty_doc_by_business_id: dict[str, dict[str, Any]] = {}
    for row in payload.faculties:
        university = university_doc_by_business_id[normalize_code(row["university_id"])]
        document = {
            "faculty_id": normalize_code(row["faculty_id"]),
            "faculty_code": coalesce_code(row["faculty_code"]),
            "faculty_name": normalize_text(row["faculty_name"]),
            "name": normalize_text(row["faculty_name"]),
            "code": coalesce_code(row["faculty_code"]),
            "university_id": str(university["_id"]),
            "university_master_id": university["university_id"],
            "university_name": university["university_name"],
            "university_code": university["university_id"],
            "is_active": True,
            "created_at": now,
            "schema_version": FACULTY_SCHEMA_VERSION,
        }
        result = await db.faculties.insert_one(document)
        faculty_doc_by_business_id[document["faculty_id"]] = {**document, "_id": result.inserted_id}

    department_doc_by_business_id: dict[str, dict[str, Any]] = {}
    for row in payload.departments:
        faculty = faculty_doc_by_business_id[normalize_code(row["faculty_id"])]
        document = {
            "department_id": normalize_code(row["department_id"]),
            "department_code": coalesce_code(row["department_code"]),
            "department_name": normalize_text(row["department_name"]),
            "name": normalize_text(row["department_name"]),
            "code": coalesce_code(row["department_code"]),
            "faculty_id": str(faculty["_id"]),
            "faculty_master_id": faculty["faculty_id"],
            "faculty_code": faculty["faculty_code"],
            "faculty_name": faculty["faculty_name"],
            "university_master_id": faculty["university_master_id"],
            "university_name": faculty["university_name"],
            "university_code": faculty["university_code"],
            "is_active": True,
            "created_at": now,
            "schema_version": DEPARTMENT_SCHEMA_VERSION,
        }
        result = await db.departments.insert_one(document)
        department_doc_by_business_id[document["department_id"]] = {**document, "_id": result.inserted_id}

    program_doc_by_business_id: dict[str, dict[str, Any]] = {}
    for row in payload.programs:
        department = department_doc_by_business_id[normalize_code(row["department_id"])]
        duration_years, total_semesters = validate_duration_and_semesters(row["duration_years"], row["total_semesters"])
        document = {
            "program_id": normalize_code(row["program_id"]),
            "program_code": coalesce_code(row["program_code"]),
            "program_name": normalize_text(row["program_name"]),
            "name": normalize_text(row["program_name"]),
            "code": coalesce_code(row["program_code"]),
            "department_id": str(department["_id"]),
            "department_master_id": department["department_id"],
            "department_name": department["department_name"],
            "department_code": department["department_code"],
            "faculty_master_id": department["faculty_master_id"],
            "faculty_code": department["faculty_code"],
            "duration_years": duration_years,
            "total_semesters": total_semesters,
            "degree_type": normalize_text(row.get("degree_type")),
            "description": None,
            "is_active": True,
            "created_at": now,
            "schema_version": PROGRAM_SCHEMA_VERSION,
        }
        result = await db.programs.insert_one(document)
        program_doc_by_business_id[document["program_id"]] = {**document, "_id": result.inserted_id}

    inserted_specializations = 0
    for row in payload.specializations:
        program_row = next(
            item
            for item in payload.programs
            if normalize_code(item["department_id"]) == normalize_code(row["department_id"])
            and coalesce_code(item["program_code"]) == coalesce_code(row["program_code"])
        )
        program = program_doc_by_business_id[normalize_code(program_row["program_id"])]
        document = {
            "specialization_id": normalize_code(row["specialization_id"]),
            "specialization_code": coalesce_code(row["specialization_code"]),
            "specialization_name": normalize_text(row["specialization_name"]),
            "name": normalize_text(row["specialization_name"]),
            "code": coalesce_code(row["specialization_code"]),
            "program_id": str(program["_id"]),
            "program_master_id": program["program_id"],
            "program_name": program["program_name"],
            "program_code": program["program_code"],
            "department_master_id": program["department_master_id"],
            "department_code": program["department_code"],
            "faculty_master_id": program["faculty_master_id"],
            "faculty_code": program["faculty_code"],
            "description": None,
            "is_active": True,
            "created_at": now,
            "schema_version": SPECIALIZATION_SCHEMA_VERSION,
        }
        await db.specializations.insert_one(document)
        inserted_specializations += 1

    return {
        "deleted_master_collections": list(MASTER_COLLECTIONS),
        "imported_counts": {
            "universities": len(university_doc_by_business_id),
            "faculties": len(faculty_doc_by_business_id),
            "departments": len(department_doc_by_business_id),
            "programs": len(program_doc_by_business_id),
            "specializations": inserted_specializations,
        },
    }


def _duplicates_by(rows: list[dict[str, Any]], field: str) -> list[dict[str, Any]]:
    counter = Counter(str(row.get(field) or "") for row in rows if row.get(field))
    return [{"value": value, "count": count} for value, count in counter.items() if count > 1]


def _scoped_duplicates_by(rows: list[dict[str, Any]], scope_fields: tuple[str, ...], field: str) -> list[dict[str, Any]]:
    counter = Counter(
        tuple(str(row.get(item) or "") for item in (*scope_fields, field))
        for row in rows
        if row.get(field)
    )
    return [
        {"scope": {name: value for name, value in zip((*scope_fields, field), values)}, "count": count}
        for values, count in counter.items()
        if count > 1
    ]


async def _fetch_collection_rows(collection_name: str, projection: dict[str, int]) -> list[dict[str, Any]]:
    collection = getattr(db, collection_name)
    return await collection.find({}, projection).to_list(length=10000)


async def post_import_audit() -> dict[str, Any]:
    universities = await _fetch_collection_rows("universities", {"university_id": 1, "university_name": 1})
    faculties = await _fetch_collection_rows("faculties", {"faculty_id": 1, "faculty_code": 1, "faculty_name": 1, "university_id": 1, "university_master_id": 1, "university_name": 1})
    departments = await _fetch_collection_rows("departments", {"department_id": 1, "department_code": 1, "department_name": 1, "faculty_id": 1, "faculty_master_id": 1, "faculty_code": 1, "faculty_name": 1, "university_master_id": 1, "university_name": 1})
    programs = await _fetch_collection_rows("programs", {"program_id": 1, "program_code": 1, "program_name": 1, "department_id": 1, "department_master_id": 1, "department_code": 1, "department_name": 1, "duration_years": 1, "total_semesters": 1, "degree_type": 1})
    specializations = await _fetch_collection_rows("specializations", {"specialization_id": 1, "specialization_code": 1, "specialization_name": 1, "program_id": 1, "program_master_id": 1, "program_code": 1, "program_name": 1})

    university_by_id = {str(row["_id"]): row for row in universities}
    faculty_by_id = {str(row["_id"]): row for row in faculties}
    department_by_id = {str(row["_id"]): row for row in departments}
    program_by_id = {str(row["_id"]): row for row in programs}
    specialization_by_id = {str(row["_id"]): row for row in specializations}

    orphan_faculties = [{"faculty_id": row.get("faculty_id"), "detail": "Parent university not found."} for row in faculties if str(row.get("university_id") or "") not in university_by_id]

    orphan_departments = []
    department_mismatches = []
    for row in departments:
        faculty = faculty_by_id.get(str(row.get("faculty_id") or ""))
        if not faculty:
            orphan_departments.append({"department_id": row.get("department_id"), "detail": "Parent faculty not found."})
            continue
        if row.get("faculty_master_id") != faculty.get("faculty_id") or row.get("faculty_code") != faculty.get("faculty_code"):
            department_mismatches.append({"department_id": row.get("department_id"), "detail": "Department faculty lineage fields do not match parent faculty."})

    orphan_programs = []
    program_mismatches = []
    invalid_program_durations = []
    for row in programs:
        department = department_by_id.get(str(row.get("department_id") or ""))
        if not department:
            orphan_programs.append({"program_id": row.get("program_id"), "detail": "Parent department not found."})
            continue
        if row.get("department_master_id") != department.get("department_id") or row.get("department_code") != department.get("department_code"):
            program_mismatches.append({"program_id": row.get("program_id"), "detail": "Program department lineage fields do not match parent department."})
        try:
            validate_duration_and_semesters(row.get("duration_years"), row.get("total_semesters"))
        except ValueError as exc:
            invalid_program_durations.append({"program_id": row.get("program_id"), "detail": str(exc)})

    orphan_specializations = []
    specialization_mismatches = []
    for row in specializations:
        program = program_by_id.get(str(row.get("program_id") or ""))
        if not program:
            orphan_specializations.append({"specialization_id": row.get("specialization_id"), "detail": "Parent program not found."})
            continue
        if row.get("program_master_id") != program.get("program_id") or row.get("program_code") != program.get("program_code"):
            specialization_mismatches.append({"specialization_id": row.get("specialization_id"), "detail": "Specialization program lineage fields do not match parent program."})

    pattern_mismatches = defaultdict(list)
    for row in faculties:
        try:
            validate_business_identifier("faculty", row.get("faculty_id"))
        except ValueError as exc:
            pattern_mismatches["faculties"].append({"faculty_id": row.get("faculty_id"), "detail": str(exc)})
    for row in departments:
        try:
            validate_business_identifier("department", row.get("department_id"))
        except ValueError as exc:
            pattern_mismatches["departments"].append({"department_id": row.get("department_id"), "detail": str(exc)})
    for row in programs:
        try:
            validate_business_identifier("program", row.get("program_id"))
        except ValueError as exc:
            pattern_mismatches["programs"].append({"program_id": row.get("program_id"), "detail": str(exc)})
    for row in specializations:
        try:
            validate_business_identifier("specialization", row.get("specialization_id"))
        except ValueError as exc:
            pattern_mismatches["specializations"].append({"specialization_id": row.get("specialization_id"), "detail": str(exc)})

    downstream_reference_findings = []
    for collection_name, fields in (
        ("users", ("faculty_id", "department_id", "program_id", "specialization_id")),
        ("batches", ("program_id", "specialization_id")),
        ("semesters", ("program_id", "specialization_id")),
        ("classes", ("faculty_id", "department_id", "program_id", "specialization_id")),
    ):
        collection = getattr(db, collection_name, None)
        if collection is None:
            continue
        rows = await collection.find({}, {field: 1 for field in fields}).to_list(length=10000)
        for row in rows:
            if row.get("faculty_id") and str(row["faculty_id"]) not in faculty_by_id:
                downstream_reference_findings.append({"collection": collection_name, "field": "faculty_id", "record_id": str(row["_id"])})
            if row.get("department_id") and str(row["department_id"]) not in department_by_id:
                downstream_reference_findings.append({"collection": collection_name, "field": "department_id", "record_id": str(row["_id"])})
            if row.get("program_id") and str(row["program_id"]) not in program_by_id:
                downstream_reference_findings.append({"collection": collection_name, "field": "program_id", "record_id": str(row["_id"])})
            if row.get("specialization_id") and str(row["specialization_id"]) not in specialization_by_id:
                downstream_reference_findings.append({"collection": collection_name, "field": "specialization_id", "record_id": str(row["_id"])})

    return {
        "master_counts": {
            "universities": len(universities),
            "faculties": len(faculties),
            "departments": len(departments),
            "programs": len(programs),
            "specializations": len(specializations),
        },
        "duplicates": {
            "university_ids": _duplicates_by(universities, "university_id"),
            "faculty_ids": _duplicates_by(faculties, "faculty_id"),
            "faculty_codes": _duplicates_by(faculties, "faculty_code"),
            "department_ids": _duplicates_by(departments, "department_id"),
            "department_codes": _scoped_duplicates_by(departments, ("faculty_id",), "department_code"),
            "program_ids": _duplicates_by(programs, "program_id"),
            "program_codes": _scoped_duplicates_by(programs, ("department_id",), "program_code"),
            "specialization_ids": _duplicates_by(specializations, "specialization_id"),
            "specialization_codes": _scoped_duplicates_by(specializations, ("program_id",), "specialization_code"),
        },
        "orphans": {
            "faculties": orphan_faculties,
            "departments": orphan_departments,
            "programs": orphan_programs,
            "specializations": orphan_specializations,
        },
        "mismatches": {
            "department_faculty": department_mismatches,
            "program_department": program_mismatches,
            "specialization_program": specialization_mismatches,
            "program_duration": invalid_program_durations,
            "patterns": dict(pattern_mismatches),
        },
        "downstream_reference_findings": downstream_reference_findings,
    }


def build_summary_markdown(
    *,
    workbook_report: dict[str, Any],
    import_report: dict[str, Any] | None,
    audit_report: dict[str, Any],
    blockers: list[dict[str, Any]],
    change_plan: dict[str, Any],
    backup_path: str | None = None,
    dry_run: bool = False,
) -> str:
    counts = audit_report["master_counts"]
    duplicates = audit_report["duplicates"]
    orphans = audit_report["orphans"]
    mismatches = audit_report["mismatches"]

    lines = [
        "# Master Hierarchy Migration Summary",
        "",
        "## Schema Changes",
        "- Added `universities` as a first-class master collection and API.",
        "- Added canonical workbook business fields to master collections while preserving legacy `name` / `code` aliases for compatibility.",
        "- Added master hierarchy indexes for business IDs, business codes, and scoped uniqueness checks.",
        "- Preserved internal ObjectId references for downstream operational entities such as batches, semesters, sections, groups, and course offerings.",
        "",
        "## Change Plan",
        f"- Dry run: {'Yes' if dry_run else 'No'}",
        *[
            f"- {collection}: add {plan['added']['count']}, update {plan['updated']['count']}, remove {plan['removed']['count']}, unchanged {plan['unchanged_count']}"
            for collection, plan in change_plan.items()
        ],
        "",
        "## Data Migration Summary",
        f"- Workbook counts validated: {json.dumps(workbook_report['sheet_counts'])}",
        f"- Workbook reconciliations applied: {len(workbook_report.get('reconciliations') or [])}",
        f"- Master collections replaced: {', '.join(import_report['deleted_master_collections']) if import_report else 'not applied'}",
        f"- Imported counts: {json.dumps(import_report['imported_counts']) if import_report else 'not applied'}",
        f"- Backup export: {backup_path or 'not requested'}",
        "",
        "## Compatibility Summary",
        "- Downstream operational collections were preserved.",
        "- Replacement is blocked when downstream collections still reference old master ObjectIds.",
        f"- Replacement blockers found during this run: {len(blockers)}",
        f"- Downstream invalid references after import: {len(audit_report['downstream_reference_findings'])}",
        "",
        "## Post-Import Audit",
        f"- Master counts: {json.dumps(counts)}",
        f"- Duplicate findings total: {sum(len(value) for value in duplicates.values())}",
        f"- Orphan findings total: {sum(len(value) for value in orphans.values())}",
        f"- Mismatch findings total: {len(mismatches['department_faculty']) + len(mismatches['program_department']) + len(mismatches['specialization_program']) + len(mismatches['program_duration']) + sum(len(value) for value in mismatches['patterns'].values())}",
        "",
        "## Assumptions",
        "- `exports/Master_copy.xlsx` is the source of truth for the core academic master hierarchy only.",
        "- Operational entities such as batches, semesters, sections, groups, course offerings, staff assignments, and student mappings are intentionally not imported from the workbook.",
        "- Existing downstream records are preserved and must continue to reference valid master ObjectIds; the script aborts instead of orphaning them.",
    ]
    return "\n".join(lines) + "\n"


async def run_import(
    workbook_path: Path = WORKBOOK_PATH,
    *,
    dry_run: bool = False,
    write_summary: bool = True,
    summary_path: Path = SUMMARY_PATH,
    backup_dir: Path | None = None,
) -> dict[str, Any]:
    payload = load_workbook_payload(workbook_path)
    workbook_report = validate_workbook_payload(payload)
    await ensure_indexes()
    current_state = await fetch_current_master_state()
    change_plan = build_master_change_plan(current_state=current_state, payload=payload)
    blockers = await detect_downstream_blockers()
    backup_path: str | None = None
    import_report: dict[str, Any] | None = None
    if dry_run:
        audit_report = await post_import_audit()
    else:
        if blockers:
            raise RuntimeError(format_downstream_blockers(blockers))
        if backup_dir is not None:
            backup_path = str(write_master_backup(current_state=current_state, backup_root=backup_dir))
        import_report = await replace_master_data(payload)
        audit_report = await post_import_audit()
    if write_summary:
        summary_path.parent.mkdir(parents=True, exist_ok=True)
        summary_path.write_text(
            build_summary_markdown(
                workbook_report=workbook_report,
                import_report=import_report,
                audit_report=audit_report,
                blockers=blockers,
                change_plan=change_plan,
                backup_path=backup_path,
                dry_run=dry_run,
            ),
            encoding="utf-8",
        )
    return {
        "workbook": workbook_report,
        "change_plan": change_plan,
        "import": import_report,
        "audit": audit_report,
        "blockers": blockers,
        "backup_path": backup_path,
        "dry_run": dry_run,
        "summary_path": str(summary_path) if write_summary else None,
    }


async def main() -> None:
    parser = argparse.ArgumentParser(description="Validate and optionally apply the CAPS AI master hierarchy workbook import.")
    parser.add_argument("--workbook", default=str(WORKBOOK_PATH), help="Path to the workbook to import.")
    parser.add_argument("--dry-run", action="store_true", help="Validate the workbook and show the change plan without mutating master data.")
    parser.add_argument("--no-summary", action="store_true", help="Skip writing the markdown summary file.")
    parser.add_argument("--summary-path", default=str(SUMMARY_PATH), help="Where to write the markdown summary.")
    parser.add_argument(
        "--backup-dir",
        default=None,
        help="Optional directory where the current master hierarchy JSON backup should be written before replacement.",
    )
    parser.add_argument(
        "--fail-on-change-plan",
        action="store_true",
        help="Exit with status 1 when the computed change plan includes adds, updates, removals, or blockers.",
    )
    args = parser.parse_args()

    result = await run_import(
        workbook_path=Path(args.workbook),
        dry_run=bool(args.dry_run),
        write_summary=not bool(args.no_summary),
        summary_path=Path(args.summary_path),
        backup_dir=Path(args.backup_dir) if args.backup_dir else None,
    )
    print(json.dumps(result, indent=2, default=str))
    if args.fail_on_change_plan and (change_plan_has_mutations(result["change_plan"]) or result["blockers"]):
        raise SystemExit(1)


if __name__ == "__main__":
    asyncio.run(main())
